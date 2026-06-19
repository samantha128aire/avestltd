#!/usr/bin/env python3
"""
SPX Double Calendar Position Monitor
Checks open positions against exit rules and color-codes column A.
Sends iMessage alert if any position is YELLOW or RED.
"""

import subprocess
import json
import sys
from datetime import datetime, date, timedelta
import urllib.request

EXCEL_PATH = '/Users/sam/Library/CloudStorage/OneDrive-Personal/Documents/SPX Calendar Analysis.xlsx'
SHEET_NAME = 'Theta Trades'

# ── Fetch live SPX and VIX ────────────────────────────────────────────────────
def get_live_data():
    """Fetch live SPX price and VIX from Yahoo Finance."""
    def fetch_yahoo(symbol):
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1m&range=1d"
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        try:
            with urllib.request.urlopen(req, timeout=10) as r:
                data = json.loads(r.read())
            meta = data['chart']['result'][0]['meta']
            return meta.get('regularMarketPrice') or meta.get('previousClose')
        except Exception as e:
            print(f"  Warning: Could not fetch {symbol}: {e}")
            return None

    spx = fetch_yahoo('%5EGSPC')
    vix = fetch_yahoo('%5EVIX')
    return spx, vix

# ── Trading day counter ───────────────────────────────────────────────────────
def count_trading_days(start_date, end_date):
    """Count trading days (Mon-Fri, no holidays) between two dates inclusive."""
    # Simple Mon-Fri count (good enough for our purposes)
    count = 0
    current = start_date
    while current <= end_date:
        if current.weekday() < 5:  # Mon=0 ... Fri=4
            count += 1
        current += timedelta(days=1)
    return count

# ── Evaluate a single position ────────────────────────────────────────────────
def evaluate_position(trade_num, entry_date, entry_price, entry_vix,
                       put_long, call_long, short_exp, long_exp,
                       spx_now, vix_now):
    """
    Apply exit rules. Returns (status, color, reasons).
    status: 'GREEN' | 'YELLOW' | 'RED'
    """
    today = date.today()
    reasons = []
    red_triggers = []
    yellow_triggers = []

    # Trading days elapsed (Day 1 = entry day)
    if isinstance(entry_date, datetime):
        entry_d = entry_date.date()
    else:
        entry_d = entry_date
    trading_days = count_trading_days(entry_d, today)

    # Short expiry days remaining
    if isinstance(short_exp, datetime):
        short_exp_d = short_exp.date()
    else:
        short_exp_d = short_exp
    days_to_short_exp = (short_exp_d - today).days

    # Drift from center
    center = entry_price
    drift = abs(spx_now - center)

    # ── RED conditions (exit immediately) ─────────────────────────────────────
    if drift > 180:
        red_triggers.append(f"DRIFT {drift:.0f} pts > 180 pt hard stop")

    if vix_now > 25:
        red_triggers.append(f"VIX {vix_now:.1f} > 25 hard stop")

    if days_to_short_exp <= 0:
        red_triggers.append(f"Short legs EXPIRED (exp {short_exp_d}) — close immediately")
    elif days_to_short_exp == 1:
        red_triggers.append(f"Short legs expire TOMORROW ({short_exp_d}) — MUST close today")
    elif days_to_short_exp == 2:
        red_triggers.append(f"Short legs expire in 2 days ({short_exp_d}) — close by end of tomorrow")

    # ── Tightened stop for Week 2 (trading days 7-11) ─────────────────────────
    if trading_days >= 7:
        if drift > 120:
            red_triggers.append(f"Week 2 drift {drift:.0f} pts > 120 pt tightened stop")

    # ── YELLOW conditions (caution) ───────────────────────────────────────────
    if not red_triggers:
        if drift > 120 and trading_days < 7:
            yellow_triggers.append(f"Drift {drift:.0f} pts — approaching 180 pt hard stop")
        elif drift > 75 and trading_days >= 7:
            yellow_triggers.append(f"Week 2 drift {drift:.0f} pts — watch closely (stop at 120)")
        elif drift > 60:
            yellow_triggers.append(f"Drift {drift:.0f} pts — above Week 1 Friday threshold (60 pts)")

        if vix_now > 22:
            yellow_triggers.append(f"VIX {vix_now:.1f} — elevated, approaching 25 hard stop")
        elif vix_now > 19:
            yellow_triggers.append(f"VIX {vix_now:.1f} — above entry filter level")

        if days_to_short_exp == 3:
            yellow_triggers.append(f"Short legs expire in 3 days ({short_exp_d}) — plan your exit")

        if trading_days == 6:
            yellow_triggers.append(f"Day 6 (Week 1 Friday checkpoint) — evaluate hold vs exit")

    # ── Determine final status ────────────────────────────────────────────────
    if red_triggers:
        status = 'RED'
        reasons = red_triggers
    elif yellow_triggers:
        status = 'YELLOW'
        reasons = yellow_triggers
    else:
        status = 'GREEN'
        reasons = [
            f"Drift {drift:.0f} pts (ok)",
            f"VIX {vix_now:.1f} (ok)",
            f"Trading day {trading_days}",
            f"Short exp in {days_to_short_exp} days"
        ]

    return status, reasons, trading_days, drift

# ── Apply color to Excel cell ─────────────────────────────────────────────────
def apply_color(cell, status):
    from openpyxl.styles import PatternFill
    colors = {
        'GREEN':  'FF92D050',  # Excel green
        'YELLOW': 'FFFFC000',  # Excel amber
        'RED':    'FFFF0000',  # Excel red
    }
    fill = PatternFill(start_color=colors[status], end_color=colors[status], fill_type='solid')
    cell.fill = fill
    cell.value = status

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    import openpyxl

    print(f"\n{'='*60}")
    print(f"SPX Monitor — {datetime.now().strftime('%Y-%m-%d %H:%M:%S CST')}")
    print(f"{'='*60}")

    # Fetch live data
    print("\nFetching live SPX + VIX...")
    spx_now, vix_now = get_live_data()
    if spx_now is None or vix_now is None:
        print("ERROR: Could not fetch live market data. Aborting.")
        sys.exit(1)
    print(f"  SPX: {spx_now:.2f}  |  VIX: {vix_now:.2f}")

    # Load workbook
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    alerts = []  # (trade_num, status, reasons)

    # Scan for open positions (rows with trade# in col B, 'buy' in col E, no sell row filled)
    row = 3
    while row <= ws.max_row:
        trade_num = ws.cell(row=row, column=2).value   # col B
        e_val     = ws.cell(row=row, column=5).value   # col E (buy/sell)
        entry_date = ws.cell(row=row, column=3).value  # col C
        entry_price = ws.cell(row=row, column=7).value # col G
        entry_vix   = ws.cell(row=row, column=8).value # col H
        put_long    = ws.cell(row=row, column=11).value # col K
        call_long   = ws.cell(row=row, column=12).value # col L
        short_exp   = ws.cell(row=row, column=13).value # col M
        long_exp    = ws.cell(row=row, column=16).value # col P

        # Only process rows that are actual open trades
        if (trade_num and isinstance(trade_num, int) and
                e_val == 'buy' and
                entry_date and entry_price):

            # Check if there's a sell row right after (position closed)
            sell_e = ws.cell(row=row+1, column=5).value
            sell_price_entry = ws.cell(row=row+1, column=7).value
            is_closed = (sell_e == 'sell' and sell_price_entry is not None)

            if is_closed:
                print(f"\n  Trade #{trade_num}: CLOSED (skipping)")
                row += 2
                continue

            status, reasons, tdays, drift = evaluate_position(
                trade_num, entry_date, entry_price, entry_vix,
                put_long, call_long, short_exp, long_exp,
                spx_now, vix_now
            )

            print(f"\n  Trade #{trade_num} (entered {entry_date.strftime('%b %d') if hasattr(entry_date,'strftime') else entry_date} @ {entry_price}, center {entry_price}):")
            print(f"    SPX now: {spx_now:.0f}  |  Drift: {drift:.0f} pts  |  VIX: {vix_now:.1f}  |  Trading day: {tdays}")
            print(f"    Status: {status}")
            for r in reasons:
                print(f"      • {r}")

            # Write status + color to col A
            status_cell = ws.cell(row=row, column=1)
            apply_color(status_cell, status)

            if status in ('YELLOW', 'RED'):
                alerts.append((trade_num, entry_date, entry_price, status, reasons, tdays, drift))

        row += 1

    # Save workbook
    wb.save(EXCEL_PATH)
    print(f"\nExcel updated: {EXCEL_PATH}")

    # Send iMessage if any alerts
    if alerts:
        msg_lines = [f"🚨 SPX MONITOR ALERT — {datetime.now().strftime('%b %d, %I:%M %p CST')}\n"]
        msg_lines.append(f"SPX: {spx_now:.2f}  |  VIX: {vix_now:.2f}\n")

        for trade_num, entry_date, entry_price, status, reasons, tdays, drift in alerts:
            icon = "🔴" if status == "RED" else "🟡"
            ed = entry_date.strftime('%b %d') if hasattr(entry_date, 'strftime') else str(entry_date)
            msg_lines.append(f"{icon} Trade #{trade_num} (entered {ed} @ {entry_price})")
            msg_lines.append(f"   Status: {status}  |  Drift: {drift:.0f} pts  |  Day {tdays}")
            for r in reasons:
                msg_lines.append(f"   • {r}")
            msg_lines.append("")

        if any(s == 'RED' for _, _, _, s, _, _, _ in alerts):
            msg_lines.append("⚠️  RED = EXIT NOW per strategy rules.")
        if any(s == 'YELLOW' for _, _, _, s, _, _, _ in alerts):
            msg_lines.append("⚠️  YELLOW = Review position carefully.")

        full_msg = "\n".join(msg_lines)
        print(f"\nSending iMessage alert...")
        print(full_msg)

        result = subprocess.run(
            ['imsg', 'send', '--to', '+15089229086', '--text', full_msg],
            capture_output=True, text=True
        )
        if result.returncode == 0:
            print("iMessage sent ✅")
        else:
            print(f"iMessage failed: {result.stderr}")
    else:
        print("\nAll positions GREEN — no alert needed.")

    print(f"\n{'='*60}\n")

if __name__ == '__main__':
    main()
